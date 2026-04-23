#!/usr/bin/env python3
"""
Sri Lanka Treasury Bond Yield Scraper

Extracts daily bond yields from treasury.gov.lk and creates formatted Excel reports.
Supports incremental updates, data validation, and anomaly detection.

Usage:
    python extract_data.py --year 2025 --month 12        # Full month export
    python extract_data.py --today                       # Update with today's report
    python extract_data.py --incremental                 # Append latest day to existing Excel
"""

from __future__ import annotations

import argparse
import base64
import json
import logging
import os
import re
import subprocess
import sys
import tempfile
import time
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from urllib.error import URLError
from urllib.request import urlopen

import openpyxl
import requests
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ───────────────────────────────────────────────────────────────────────────────
# Configuration
# ───────────────────────────────────────────────────────────────────────────────
TREASURY_URL = "https://www.treasury.gov.lk/web/report-daily-report/section/{year}"
DEFAULT_REPORTS_DIR = Path("treasury_reports")
DEFAULT_OUTPUT_DIR = Path("output")
DEFAULT_OUTPUT_FILE = DEFAULT_OUTPUT_DIR / "treasury_bond_yields.xlsx"
MAX_RETRIES = 3
RETRY_DELAY = 5  # seconds

# ───────────────────────────────────────────────────────────────────────────────
# Logging
# ───────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ───────────────────────────────────────────────────────────────────────────────
# Helpers
# ───────────────────────────────────────────────────────────────────────────────

def excel_date_to_datetime(excel_date: float) -> Optional[datetime]:
    """Convert Excel date serial to datetime."""
    if isinstance(excel_date, (int, float)):
        return datetime(1899, 12, 30) + timedelta(days=float(excel_date))
    return None


def parse_bond_number(bond_str) -> Optional[str]:
    """Clean bond number string."""
    if not bond_str:
        return None
    return str(bond_str).strip().replace("%", "")


def retry_download(url: str, filepath: Path, max_retries: int = MAX_RETRIES) -> bool:
    """Download a file with retry logic."""
    for attempt in range(1, max_retries + 1):
        try:
            result = subprocess.run(
                ["curl", "-L", "-o", str(filepath), url],
                capture_output=True,
                timeout=60,
            )
            if result.returncode == 0 and filepath.exists() and filepath.stat().st_size > 100_000:
                logger.info(f"Downloaded: {filepath.name} ({filepath.stat().st_size:,} bytes)")
                return True
            else:
                logger.warning(f"Attempt {attempt}/{max_retries} failed for {filepath.name}")
        except Exception as e:
            logger.warning(f"Attempt {attempt}/{max_retries} error: {e}")

        if attempt < max_retries:
            time.sleep(RETRY_DELAY * attempt)

    logger.error(f"Failed to download after {max_retries} attempts: {url}")
    return False


def validate_report(filepath: Path) -> bool:
    """Validate that a downloaded report is valid and contains expected sheets."""
    if not filepath.exists():
        return False
    if filepath.stat().st_size < 100_000:
        logger.warning(f"File too small: {filepath.name}")
        return False

    try:
        wb = xlrd.open_workbook(str(filepath), ragged_rows=True)
        if "QuotesTBond" not in wb.sheet_names():
            logger.warning(f"Missing QuotesTBond sheet: {filepath.name}")
            return False
        return True
    except Exception as e:
        logger.warning(f"Invalid Excel file {filepath.name}: {e}")
        return False


# ───────────────────────────────────────────────────────────────────────────────
# Treasury Scraper
# ───────────────────────────────────────────────────────────────────────────────

class TreasuryScraper:
    """Scraper for Sri Lanka Treasury daily bond reports."""

    def __init__(self, reports_dir: Path = DEFAULT_REPORTS_DIR):
        self.reports_dir = reports_dir
        self.reports_dir.mkdir(parents=True, exist_ok=True)

    def fetch_report_list(self, year: int = 2025) -> Dict[datetime, str]:
        """Scrape treasury.gov.lk to get Daily Summary Report URLs."""
        logger.info(f"Fetching report list for {year}...")
        url = TREASURY_URL.format(year=year)

        for attempt in range(1, MAX_RETRIES + 1):
            try:
                response = urlopen(url, timeout=30)
                html = response.read().decode("utf-8")
                break
            except URLError as e:
                logger.warning(f"Attempt {attempt} failed: {e}")
                if attempt == MAX_RETRIES:
                    logger.error("Max retries exceeded fetching report list")
                    return {}
                time.sleep(RETRY_DELAY * attempt)

        reports = {}
        # Pattern to find date + UUID + "Daily Summary Report"
        pattern = r"(\d{2})\.(\d{2})\.(\d{4}).*?api/file/([a-f0-9-]+).*?Daily Summary Report"
        matches = re.findall(pattern, html, re.DOTALL)

        for match in matches:
            day, month, year_str, uuid = match
            date_str = f"{day}.{month}.{year_str}"
            try:
                report_date = datetime.strptime(date_str, "%d.%m.%Y")
                report_url = f"https://www.treasury.gov.lk/api/file/{uuid}"
                reports[report_date] = report_url
            except ValueError:
                continue

        logger.info(f"Found {len(reports)} reports for {year}")
        return reports

    def download_report(self, url: str, date: datetime) -> Optional[Path]:
        """Download a single report."""
        date_str = date.strftime("%d-%m-%Y")
        filename = f"report_{date_str}.xlsx"
        filepath = self.reports_dir / filename

        if filepath.exists() and validate_report(filepath):
            logger.debug(f"Already downloaded: {filename}")
            return filepath

        if retry_download(url, filepath):
            if validate_report(filepath):
                return filepath
            else:
                filepath.unlink(missing_ok=True)

        return None

    def extract_data_from_report(
        self, filepath: Path
    ) -> Tuple[Dict[Tuple[str, datetime], float], Dict[Tuple[str, datetime], float]]:
        """Extract QuotesTBond data from a single report."""
        data_twoway: Dict[Tuple[str, datetime], float] = {}
        data_ddo: Dict[Tuple[str, datetime], float] = {}

        try:
            wb = xlrd.open_workbook(str(filepath), ragged_rows=True)
            ws = wb.sheet_by_name("QuotesTBond")
        except Exception as e:
            logger.error(f"Cannot open {filepath.name}: {e}")
            return data_twoway, data_ddo

        # TWO WAY QUOTES section (rows 8-72)
        for row_idx in range(8, 73):
            self._parse_row(ws, row_idx, data_twoway)

        # DDO section (rows 75 onwards)
        for row_idx in range(75, ws.nrows):
            self._parse_row(ws, row_idx, data_ddo)

        return data_twoway, data_ddo

    def _parse_row(
        self,
        ws,
        row_idx: int,
        data_dict: Dict[Tuple[str, datetime], float],
    ) -> None:
        """Parse a single row and add to data dictionary."""
        try:
            bond_number = parse_bond_number(ws.cell_value(row_idx, 2))
            maturity_date_val = ws.cell_value(row_idx, 4)
            buy_yield = ws.cell_value(row_idx, 7)

            if not bond_number:
                return

            maturity_date = excel_date_to_datetime(maturity_date_val)
            if maturity_date and buy_yield and buy_yield != 0:
                key = (bond_number, maturity_date)
                data_dict[key] = float(buy_yield)
        except (IndexError, ValueError, TypeError):
            pass

    def validate_extracted_data(
        self,
        date: datetime,
        twoway: Dict,
        ddo: Dict,
    ) -> List[str]:
        """Validate extracted data and return list of issues."""
        issues = []

        if not twoway and not ddo:
            issues.append(f"No data extracted for {date.strftime('%Y-%m-%d')}")
            return issues

        if not twoway:
            issues.append(f"No TWO WAY data for {date.strftime('%Y-%m-%d')}")
        if not ddo:
            issues.append(f"No DDO data for {date.strftime('%Y-%m-%d')}")

        # Check for zero yields (potential data issue)
        zero_yields = sum(1 for v in twoway.values() if v == 0) + sum(
            1 for v in ddo.values() if v == 0
        )
        if zero_yields > 0:
            issues.append(f"{zero_yields} bonds with zero yields on {date.strftime('%Y-%m-%d')}")

        # Check for suspicious yield values (>50% or <0%)
        for data, name in [(twoway, "TWO WAY"), (ddo, "DDO")]:
            for (bond, mat_date), yield_val in data.items():
                if yield_val > 0.50:
                    issues.append(
                        f"High yield alert: {bond} = {yield_val:.2%} ({name})"
                    )
                elif yield_val < 0:
                    issues.append(
                        f"Negative yield: {bond} = {yield_val:.2%} ({name})"
                    )

        return issues


# ───────────────────────────────────────────────────────────────────────────────
# Excel Builder
# ───────────────────────────────────────────────────────────────────────────────

class ExcelBuilder:
    """Builds formatted Excel output for bond yields."""

    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BOND_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def create_workbook(
        self,
        data_twoway: Dict[Tuple[str, datetime], Dict[datetime, float]],
        data_ddo: Dict[Tuple[str, datetime], Dict[datetime, float]],
        all_dates: List[datetime],
        output_path: Path,
    ) -> Path:
        """Create formatted Excel workbook."""
        logger.info(f"Creating Excel: {output_path}")
        wb = Workbook()
        ws_twoway = wb.active
        ws_twoway.title = "TWO_WAY_QUOTES"
        ws_ddo = wb.create_sheet("DDO_EDR_BONDS")

        sorted_dates = sorted(all_dates)
        date_formats = [d.strftime("%d-%b-%y") for d in sorted_dates]

        # Write both sheets
        for ws, data_dict in [(ws_twoway, data_twoway), (ws_ddo, data_ddo)]:
            self._write_sheet(ws, data_dict, sorted_dates, date_formats)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info(f"Saved: {output_path}")
        return output_path

    def _write_sheet(
        self,
        ws,
        data_dict: Dict[Tuple[str, datetime], Dict[datetime, float]],
        sorted_dates: List[datetime],
        date_formats: List[str],
    ) -> None:
        """Write a single sheet with bond data."""
        # Header
        ws.cell(1, 1, "Bond Number")
        ws.cell(1, 2, "Maturity Date")
        for col, date_str in enumerate(date_formats, 3):
            cell = ws.cell(1, col, date_str)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.THIN_BORDER

        for col in [1, 2]:
            ws.cell(1, col).fill = self.HEADER_FILL
            ws.cell(1, col).font = self.HEADER_FONT
            ws.cell(1, col).border = self.THIN_BORDER
            ws.cell(1, col).alignment = Alignment(horizontal="center")

        # Data
        all_bonds = sorted(data_dict.keys(), key=lambda x: x[1])
        for row_idx, (bond_number, maturity_date) in enumerate(all_bonds, 2):
            ws.cell(row_idx, 1, bond_number)
            ws.cell(row_idx, 2, maturity_date.strftime("%d-%b-%Y"))

            ws.cell(row_idx, 1).fill = self.BOND_FILL
            ws.cell(row_idx, 2).fill = self.BOND_FILL

            for col in range(1, len(sorted_dates) + 3):
                ws.cell(row_idx, col).border = self.THIN_BORDER

            yields_dict = data_dict.get((bond_number, maturity_date), {})
            for col, date in enumerate(sorted_dates, 3):
                yield_val = yields_dict.get(date)
                if yield_val is not None and yield_val != 0:
                    ws.cell(row_idx, col, yield_val)
                    ws.cell(row_idx, col).number_format = "0.00%"
                    ws.cell(row_idx, col).alignment = Alignment(horizontal="right")

        # Column widths
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 14
        for col in range(3, len(sorted_dates) + 3):
            ws.column_dimensions[get_column_letter(col)].width = 12

    def append_new_date(
        self,
        existing_path: Path,
        new_data_twoway: Dict[Tuple[str, datetime], float],
        new_data_ddo: Dict[Tuple[str, datetime], float],
        new_date: datetime,
    ) -> bool:
        """Append a new date column to existing Excel (incremental update)."""
        if not existing_path.exists():
            return False

        try:
            wb = openpyxl.load_workbook(existing_path)
        except Exception as e:
            logger.error(f"Cannot open existing Excel: {e}")
            return False

        date_str = new_date.strftime("%d-%b-%y")
        logger.info(f"Appending date column: {date_str}")

        for ws_name, new_data in [("TWO_WAY_QUOTES", new_data_twoway), ("DDO_EDR_BONDS", new_data_ddo)]:
            if ws_name not in wb.sheetnames:
                continue
            ws = wb[ws_name]

            # Find next empty column
            max_col = ws.max_column + 1
            header_cell = ws.cell(1, max_col, date_str)
            header_cell.fill = self.HEADER_FILL
            header_cell.font = self.HEADER_FONT
            header_cell.alignment = Alignment(horizontal="center")
            header_cell.border = self.THIN_BORDER

            # Find bond rows and add yields
            bond_col_map = {}
            for row in range(2, ws.max_row + 1):
                bond = ws.cell(row, 1).value
                mat_date_str = ws.cell(row, 2).value
                if bond and mat_date_str:
                    try:
                        mat_date = datetime.strptime(mat_date_str, "%d-%b-%Y")
                        bond_col_map[(bond, mat_date)] = row
                    except ValueError:
                        pass

            # Add yields for matching bonds
            for (bond, mat_date), yield_val in new_data.items():
                if (bond, mat_date) in bond_col_map:
                    row = bond_col_map[(bond, mat_date)]
                    cell = ws.cell(row, max_col, yield_val)
                    cell.number_format = "0.00%"
                    cell.alignment = Alignment(horizontal="right")
                    cell.border = self.THIN_BORDER
                else:
                    # New bond - add row
                    new_row = ws.max_row + 1
                    ws.cell(new_row, 1, bond)
                    ws.cell(new_row, 2, mat_date.strftime("%d-%b-%Y"))
                    ws.cell(new_row, 1).fill = self.BOND_FILL
                    ws.cell(new_row, 2).fill = self.BOND_FILL
                    for col in range(1, max_col):
                        ws.cell(new_row, col).border = self.THIN_BORDER
                    cell = ws.cell(new_row, max_col, yield_val)
                    cell.number_format = "0.00%"
                    cell.alignment = Alignment(horizontal="right")
                    cell.border = self.THIN_BORDER

            # Adjust column width
            ws.column_dimensions[get_column_letter(max_col)].width = 12

        wb.save(existing_path)
        logger.info(f"Updated: {existing_path}")
        return True


# ───────────────────────────────────────────────────────────────────────────────
# Email Sender
# ───────────────────────────────────────────────────────────────────────────────

class EmailSender:
    """Send reports via Resend API."""

    def __init__(self, api_key: str, from_email: str, to_email: str):
        self.api_key = api_key
        self.from_email = from_email
        self.to_email = to_email

    def send_report(
        self,
        excel_path: Path,
        date: datetime,
        stats: Optional[Dict] = None,
    ) -> bool:
        """Send Excel report as email attachment with summary."""
        if not excel_path.exists():
            logger.error(f"Excel file not found: {excel_path}")
            return False

        with open(excel_path, "rb") as f:
            file_b64 = base64.b64encode(f.read()).decode()

        date_str = date.strftime("%Y-%m-%d")
        subject = f"Treasury Bond Yields - {date_str}"

        # Build HTML email with stats
        html_body = self._build_html_email(date_str, stats)

        payload = {
            "from": self.from_email,
            "to": [self.to_email],
            "subject": subject,
            "html": html_body,
            "attachments": [
                {
                    "filename": excel_path.name,
                    "content": file_b64,
                }
            ],
        }

        try:
            r = requests.post(
                "https://api.resend.com/emails",
                headers={
                    "Authorization": f"Bearer {self.api_key}",
                    "Content-Type": "application/json",
                },
                json=payload,
                timeout=30,
            )
            if r.status_code in (200, 201, 202):
                logger.info("Email sent successfully")
                return True
            else:
                logger.error(f"Email failed: {r.status_code} - {r.text}")
                return False
        except Exception as e:
            logger.error(f"Email error: {e}")
            return False

    def _build_html_email(self, date_str: str, stats: Optional[Dict]) -> str:
        """Build HTML email body with summary."""
        html = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <h2 style="color: #1F4E79;">Treasury Bond Yields Report</h2>
            <p>Date: <strong>{date_str}</strong></p>
        """

        if stats:
            html += f"""
            <h3>Summary</h3>
            <table style="border-collapse: collapse; width: 400px;">
                <tr>
                    <td style="border: 1px solid #ddd; padding: 8px;">TWO WAY Bonds</td>
                    <td style="border: 1px solid #ddd; padding: 8px; text-align: right;">{stats.get('twoway_count', 0)}</td>
                </tr>
                <tr>
                    <td style="border: 1px solid #ddd; padding: 8px;">DDO Bonds</td>
                    <td style="border: 1px solid #ddd; padding: 8px; text-align: right;">{stats.get('ddo_count', 0)}</td>
                </tr>
                <tr>
                    <td style="border: 1px solid #ddd; padding: 8px;">Total Dates</td>
                    <td style="border: 1px solid #ddd; padding: 8px; text-align: right;">{stats.get('total_dates', 0)}</td>
                </tr>
            </table>
            """

            if stats.get("issues"):
                html += "<h3>Alerts</h3><ul>"
                for issue in stats["issues"][-5:]:
                    html += f"<li>{issue}</li>"
                html += "</ul>"

        html += """
            <p>Please find the detailed report attached.</p>
            <hr>
            <p style="font-size: 12px; color: #666;">
                Generated by bond-scrapper | 
                <a href="https://github.com/BishanFer/bond-scrapper">GitHub</a>
            </p>
        </body>
        </html>
        """
        return html


# ───────────────────────────────────────────────────────────────────────────────
# Main Runner
# ───────────────────────────────────────────────────────────────────────────────

class BondScraperRunner:
    """Main runner that orchestrates scraping, building, and emailing."""

    def __init__(
        self,
        reports_dir: Path = DEFAULT_REPORTS_DIR,
        output_path: Path = DEFAULT_OUTPUT_FILE,
    ):
        self.scraper = TreasuryScraper(reports_dir)
        self.builder = ExcelBuilder()
        self.output_path = output_path

    def run_full_export(self, year: int, month: Optional[int] = None) -> Optional[Path]:
        """Run full export for a year/month."""
        reports = self.scraper.fetch_report_list(year)
        if not reports:
            logger.error("No reports found")
            return None

        if month:
            reports = {k: v for k, v in reports.items() if k.month == month}
            logger.info(f"Filtered to {len(reports)} reports for month {month}")

        # Download all
        downloaded = []
        for date, url in sorted(reports.items()):
            filepath = self.scraper.download_report(url, date)
            if filepath:
                downloaded.append((date, filepath))

        if not downloaded:
            logger.error("No reports downloaded")
            return None

        # Extract data
        data_twoway = defaultdict(dict)
        data_ddo = defaultdict(dict)
        all_dates = []
        all_issues = []

        for date, filepath in downloaded:
            twoway, ddo = self.scraper.extract_data_from_report(filepath)
            all_dates.append(date)

            for key, yield_val in twoway.items():
                data_twoway[key][date] = yield_val
            for key, yield_val in ddo.items():
                data_ddo[key][date] = yield_val

            # Validate
            issues = self.scraper.validate_extracted_data(date, twoway, ddo)
            all_issues.extend(issues)

        if issues := all_issues:
            for issue in issues:
                logger.warning(f"Validation: {issue}")

        # Build Excel
        path = self.builder.create_workbook(
            dict(data_twoway), dict(data_ddo), all_dates, self.output_path
        )

        stats = {
            "twoway_count": len(data_twoway),
            "ddo_count": len(data_ddo),
            "total_dates": len(set(all_dates)),
            "issues": all_issues,
        }

        logger.info(f"Complete: {stats['twoway_count']} TWO WAY, {stats['ddo_count']} DDO bonds")
        return path, stats

    def run_incremental(self) -> Optional[Path]:
        """Fetch today's report and append to existing Excel."""
        today = datetime.now()
        reports = self.scraper.fetch_report_list(today.year)

        # Find today's or most recent report
        target_date = None
        target_url = None
        date_str = today.strftime("%d.%m.%Y")

        for date, url in reports.items():
            if date.strftime("%d.%m.%Y") == date_str:
                target_date = date
                target_url = url
                break

        if not target_date:
            # Try yesterday
            yesterday = today - timedelta(days=1)
            date_str_yest = yesterday.strftime("%d.%m.%Y")
            for date, url in reports.items():
                if date.strftime("%d.%m.%Y") == date_str_yest:
                    target_date = date
                    target_url = url
                    logger.info(f"Using yesterday's report: {date_str_yest}")
                    break

        if not target_date:
            logger.error("No recent report found")
            return None

        # Download
        filepath = self.scraper.download_report(target_url, target_date)
        if not filepath:
            return None

        # Extract
        twoway, ddo = self.scraper.extract_data_from_report(filepath)
        logger.info(f"Extracted: {len(twoway)} TWO WAY, {len(ddo)} DDO bonds")

        # Validate
        issues = self.scraper.validate_extracted_data(target_date, twoway, ddo)
        for issue in issues:
            logger.warning(f"Validation: {issue}")

        # Append or create
        if self.output_path.exists():
            success = self.builder.append_new_date(self.output_path, twoway, ddo, target_date)
            if not success:
                logger.warning("Incremental update failed, falling back to full rebuild")
                return self.run_full_export(target_date.year, target_date.month)
        else:
            return self.run_full_export(target_date.year, target_date.month)

        stats = {
            "twoway_count": len(twoway),
            "ddo_count": len(ddo),
            "total_dates": "incremental",
            "issues": issues,
        }

        return self.output_path, stats


# ───────────────────────────────────────────────────────────────────────────────
# CLI Entry Point
# ───────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Sri Lanka Treasury Bond Yield Scraper",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python extract_data.py --year 2025 --month 12     # Full month export
  python extract_data.py --today                     # Update with latest report
  python extract_data.py --incremental               # Append latest day to existing Excel
        """,
    )
    parser.add_argument("--year", type=int, default=2025, help="Year to scrape")
    parser.add_argument("--month", type=int, help="Month to filter (1-12)")
    parser.add_argument(
        "--today", action="store_true", help="Fetch latest report only"
    )
    parser.add_argument(
        "--incremental",
        action="store_true",
        help="Append latest day to existing Excel",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_FILE,
        help="Output Excel file path",
    )

    args = parser.parse_args()

    runner = BondScraperRunner(output_path=args.output)

    if args.incremental or args.today:
        result = runner.run_incremental()
    else:
        result = runner.run_full_export(args.year, args.month)

    if result:
        path, stats = result
        print(f"\n✓ Report saved: {path}")
        print(f"  TWO WAY bonds: {stats['twoway_count']}")
        print(f"  DDO bonds: {stats['ddo_count']}")
        if stats.get("issues"):
            print(f"  ⚠ Issues found: {len(stats['issues'])}")
    else:
        print("\n✗ Failed to generate report")
        sys.exit(1)


if __name__ == "__main__":
    main()